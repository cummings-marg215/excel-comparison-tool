import re
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


MIN_WORD_LENGTH = 4

STOPWORDS = {
    "the", "and", "for", "with", "from", "that", "this",
    "are", "was", "were", "have", "has", "into", "onto",
    "shall", "must", "should", "will", "their", "there",
    "you", "your", "its", "they", "them", "than", "then",
    "also", "such", "each", "when", "where", "using", "based"
}

CONTROL_ID_PATTERN = re.compile(r"\bCORE\s*#?\s*(\d+(?:\.\d+)?)\b", re.IGNORECASE)
ISE_ID_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*$")


def normalize_ise_id(text):
    if pd.isna(text):
        return None
    match = ISE_ID_PATTERN.match(str(text).strip())
    return match.group(1) if match else None


def normalize_control_id(text):
    if pd.isna(text):
        return None
    match = CONTROL_ID_PATTERN.search(str(text))
    if not match:
        return None
    return f"CORE #{match.group(1)}".upper()


def normalize_join_key(value):
    if pd.isna(value):
        return ""
    return str(value).strip().upper()


def excel_col_to_index(col):
    col = str(col).upper().strip()
    index = 0
    for char in col:
        if not char.isalpha():
            raise ValueError(f"Invalid Excel column: {col}")
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def get_column_name(df, user_input):
    """
    Resolve either an exact header name or an Excel column letter.
    """
    user_input = str(user_input).strip()

    if not user_input:
        raise ValueError("Column input cannot be blank.")

    if user_input in df.columns:
        return user_input

    lower_map = {str(col).strip().lower(): col for col in df.columns}
    if user_input.lower() in lower_map:
        return lower_map[user_input.lower()]

    if user_input.replace(" ", "").isalpha():
        col_index = excel_col_to_index(user_input)
        if col_index >= len(df.columns):
            raise IndexError(
                f"Column {user_input} does not exist. "
                f"The selected sheet only has {len(df.columns)} columns."
            )
        return df.columns[col_index]

    raise ValueError(
        f"Could not resolve column '{user_input}'. "
        "Enter an exact column header or Excel column letter."
    )


def clean_words(text):
    if pd.isna(text):
        return []

    text = str(text).lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    words = text.split()

    words = [
        word for word in words
        if len(word) >= MIN_WORD_LENGTH and word not in STOPWORDS
    ]

    return sorted(set(words))


def compare_text(source_text, target_text):
    source_ise_id = normalize_ise_id(source_text)
    target_ise_id = normalize_ise_id(target_text)

    if source_ise_id and target_ise_id:
        if source_ise_id == target_ise_id:
            return {
                "Match %": 1,
                "Matched Terms": source_ise_id,
                "Missing Terms": "No major missing terms",
                "Match Quality": "High",
                "Notes": "Exact ISE ID match."
            }
        return {
            "Match %": 0,
            "Matched Terms": "",
            "Missing Terms": source_ise_id,
            "Match Quality": "No Match",
            "Notes": f"ISE ID mismatch: source={source_ise_id}, target={target_ise_id}."
        }

    source_control_id = normalize_control_id(source_text)
    target_control_id = normalize_control_id(target_text)

    if source_control_id and target_control_id:
        if source_control_id == target_control_id:
            return {
                "Match %": 1,
                "Matched Terms": source_control_id,
                "Missing Terms": "No major missing terms",
                "Match Quality": "High",
                "Notes": "Exact CORE control ID match."
            }
        return {
            "Match %": 0,
            "Matched Terms": "",
            "Missing Terms": source_control_id,
            "Match Quality": "No Match",
            "Notes": f"CORE control ID mismatch: source={source_control_id}, target={target_control_id}."
        }

    source_words = clean_words(source_text)
    target_words = set(clean_words(target_text))

    if not source_words:
        return {
            "Match %": 0,
            "Matched Terms": "",
            "Missing Terms": "No source terms",
            "Match Quality": "No Match",
            "Notes": "No usable source words to compare."
        }

    matched_words = [word for word in source_words if word in target_words]
    missing_words = [word for word in source_words if word not in target_words]
    match_percent = len(matched_words) / len(source_words)

    if match_percent >= 0.85:
        quality = "High"
        notes = "Strong alignment. Most key source terms appear in the target text."
    elif match_percent >= 0.60:
        quality = "Medium"
        notes = "Partial alignment. Review missing terms for possible gaps."
    elif match_percent > 0:
        quality = "Low"
        notes = "Weak alignment. Several key source terms are missing."
    else:
        quality = "No Match"
        notes = "No meaningful overlap found."

    return {
        "Match %": match_percent,
        "Matched Terms": ", ".join(matched_words),
        "Missing Terms": ", ".join(missing_words) if missing_words else "No major missing terms",
        "Match Quality": quality,
        "Notes": notes
    }


def safe_sheet_name(name):
    invalid_chars = r'[]:*?/\\'
    name = str(name)
    for char in invalid_chars:
        name = name.replace(char, "-")
    return name[:31]


def parse_sheet_value(value):
    value = str(value).strip()
    if value.isdigit():
        return int(value)
    return value


def load_excel_dataframe(uploaded_file, sheet_name):
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name)


def get_excel_sheet_names(uploaded_file):
    uploaded_file.seek(0)
    xls = pd.ExcelFile(uploaded_file)
    return xls.sheet_names


def run_row_by_row_comparison(
    name,
    source_file,
    source_sheet,
    source_col,
    target_file,
    target_sheet,
    target_col
):
    source_df = load_excel_dataframe(source_file, source_sheet)
    target_df = load_excel_dataframe(target_file, target_sheet)

    source_column_name = get_column_name(source_df, source_col)
    target_column_name = get_column_name(target_df, target_col)

    source_series = source_df[source_column_name]
    target_series = target_df[target_column_name]

    max_rows = max(len(source_series), len(target_series))
    results = []

    for i in range(max_rows):
        source_text = source_series.iloc[i] if i < len(source_series) else ""
        target_text = target_series.iloc[i] if i < len(target_series) else ""
        comparison = compare_text(source_text, target_text)

        results.append({
            "Source Row": i + 2,
            "Target Row": i + 2,
            "Comparison Name": name,
            "Match Mode": "Row by Row",
            "Join Key": "",
            "Source File": source_file.name,
            "Source Sheet": source_sheet,
            "Source Column": source_column_name,
            "Source Text": source_text,
            "Target File": target_file.name,
            "Target Sheet": target_sheet,
            "Target Column": target_column_name,
            "Target Text": target_text,
            **comparison
        })

    return pd.DataFrame(results)


def run_key_based_comparison(
    name,
    source_file,
    source_sheet,
    source_key_col,
    source_compare_col,
    target_file,
    target_sheet,
    target_key_col,
    target_compare_col,
    join_type
):
    source_df = load_excel_dataframe(source_file, source_sheet)
    target_df = load_excel_dataframe(target_file, target_sheet)

    source_key_name = get_column_name(source_df, source_key_col)
    source_compare_name = get_column_name(source_df, source_compare_col)
    target_key_name = get_column_name(target_df, target_key_col)
    target_compare_name = get_column_name(target_df, target_compare_col)

    source_df = source_df.copy()
    target_df = target_df.copy()

    source_df["__source_row__"] = source_df.index + 2
    target_df["__target_row__"] = target_df.index + 2

    source_df["__join_key__"] = source_df[source_key_name].apply(normalize_join_key)
    target_df["__join_key__"] = target_df[target_key_name].apply(normalize_join_key)

    source_df = source_df[source_df["__join_key__"] != ""]
    target_df = target_df[target_df["__join_key__"] != ""]

    source_duplicates = set(
        source_df.loc[source_df["__join_key__"].duplicated(keep=False), "__join_key__"]
    )
    target_duplicates = set(
        target_df.loc[target_df["__join_key__"].duplicated(keep=False), "__join_key__"]
    )

    merged = pd.merge(
        source_df[["__join_key__", "__source_row__", source_key_name, source_compare_name]],
        target_df[["__join_key__", "__target_row__", target_key_name, target_compare_name]],
        on="__join_key__",
        how=join_type,
        suffixes=("_source", "_target")
    )

    results = []

    for _, row in merged.iterrows():
        join_key = row.get("__join_key__", "")
        source_text = row.get(source_compare_name, "")
        target_text = row.get(target_compare_name, "")

        source_missing = pd.isna(source_text)
        target_missing = pd.isna(target_text)

        if source_missing and not target_missing:
            comparison = {
                "Match %": 0,
                "Matched Terms": "",
                "Missing Terms": "Missing source record/text",
                "Match Quality": "No Match",
                "Notes": "Join key exists in target, but source comparison text is missing."
            }
        elif target_missing and not source_missing:
            comparison = {
                "Match %": 0,
                "Matched Terms": "",
                "Missing Terms": "Missing target record/text",
                "Match Quality": "No Match",
                "Notes": "Join key exists in source, but target comparison text is missing."
            }
        elif source_missing and target_missing:
            comparison = {
                "Match %": 0,
                "Matched Terms": "",
                "Missing Terms": "Missing source and target text",
                "Match Quality": "No Match",
                "Notes": "Both source and target comparison text are missing."
            }
        else:
            comparison = compare_text(source_text, target_text)

        duplicate_notes = []
        if join_key in source_duplicates:
            duplicate_notes.append("Duplicate source key detected.")
        if join_key in target_duplicates:
            duplicate_notes.append("Duplicate target key detected.")
        if duplicate_notes:
            comparison["Notes"] = comparison["Notes"] + " " + " ".join(duplicate_notes)

        results.append({
            "Source Row": row.get("__source_row__", ""),
            "Target Row": row.get("__target_row__", ""),
            "Comparison Name": name,
            "Match Mode": f"Key Based ({join_type})",
            "Join Key": join_key,
            "Source File": source_file.name,
            "Source Sheet": source_sheet,
            "Source Key Column": source_key_name,
            "Source Compare Column": source_compare_name,
            "Source Text": "" if pd.isna(source_text) else source_text,
            "Target File": target_file.name,
            "Target Sheet": target_sheet,
            "Target Key Column": target_key_name,
            "Target Compare Column": target_compare_name,
            "Target Text": "" if pd.isna(target_text) else target_text,
            **comparison
        })

    return pd.DataFrame(results)


def build_summary(all_results):
    summary_rows = []

    for comparison_name, group in all_results.groupby("Comparison Name"):
        summary_rows.append({
            "Comparison Name": comparison_name,
            "Total Rows Compared": len(group),
            "High Matches": (group["Match Quality"] == "High").sum(),
            "Medium Matches": (group["Match Quality"] == "Medium").sum(),
            "Low Matches": (group["Match Quality"] == "Low").sum(),
            "No Matches": (group["Match Quality"] == "No Match").sum(),
            "Rows Needing Review": (group["Match Quality"] != "High").sum(),
            "Average Match %": group["Match %"].mean()
        })

    return pd.DataFrame(summary_rows)


def format_workbook_openpyxl(output_buffer):
    output_buffer.seek(0)
    wb = load_workbook(output_buffer)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            header = ws.cell(row=1, column=col).value
            if header and any(term in str(header).lower() for term in ["text", "terms", "notes"]):
                ws.column_dimensions[col_letter].width = 45
            elif header and "match %" in str(header).lower():
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 18

        headers = [cell.value for cell in ws[1]]
        if "Match %" in headers:
            match_col = headers.index("Match %") + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=match_col).number_format = "0.00%"

    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
    return final_buffer


def create_output_workbook(results_list):
    all_results = pd.concat(results_list, ignore_index=True)
    summary = build_summary(all_results)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")

        for df in results_list:
            comparison_name = df["Comparison Name"].iloc[0]
            sheet_name = safe_sheet_name(comparison_name)
            df.to_excel(writer, index=False, sheet_name=sheet_name)

    return format_workbook_openpyxl(output), summary, all_results


st.set_page_config(
    page_title="Excel Column Comparison Tool",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Excel Column Comparison Tool")

st.write(
    "Upload two Excel files, choose whether to compare row-by-row or join records by a shared key "
    "such as canonical_item_id, then download a clean results workbook."
)

with st.expander("How the score is calculated", expanded=False):
    st.write(
        """
        **Recommended mode: Key-based comparison**

        This joins the two uploaded files using a shared ID such as `canonical_item_id`.
        Example: `ISE 01.CP01` can match even if it is row 74 in one file and row 10 in another.

        Then the tool compares the selected source and target text columns.

        For text comparison, it:
        - lowercases the text
        - removes punctuation
        - removes common stopwords
        - removes duplicate terms
        - compares source terms against target terms
        - calculates: matched source terms / total source terms

        It also handles exact ISE ID and CORE ID matches.
        """
    )

st.subheader("1. Upload Files")

col1, col2 = st.columns(2)

with col1:
    source_file = st.file_uploader(
        "Upload Source Excel File",
        type=["xlsx", "xls"],
        key="source_file"
    )

with col2:
    target_file = st.file_uploader(
        "Upload Target Excel File",
        type=["xlsx", "xls"],
        key="target_file"
    )

st.subheader("2. Choose Comparison Mode")

comparison_mode = st.radio(
    "Comparison Mode",
    options=[
        "Key-based join using canonical_item_id or another shared ID",
        "Row-by-row comparison"
    ],
    index=0
)

st.subheader("3. Configure Comparison")

comparison_name = st.text_input("Comparison Name", value="Comparison 1")

if source_file is not None and target_file is not None:
    try:
        source_sheets = get_excel_sheet_names(source_file)
        target_sheets = get_excel_sheet_names(target_file)

        source_sheet = st.selectbox("Source Sheet", source_sheets, index=0)
        target_sheet = st.selectbox("Target Sheet", target_sheets, index=0)

        source_preview = load_excel_dataframe(source_file, source_sheet)
        target_preview = load_excel_dataframe(target_file, target_sheet)

        source_columns = list(source_preview.columns)
        target_columns = list(target_preview.columns)

        st.caption(
            "Tip: For ISE/RISO files, select `canonical_item_id` as the join key when available."
        )

        if comparison_mode.startswith("Key-based"):
            col1, col2 = st.columns(2)

            with col1:
                source_key_col = st.selectbox(
                    "Source Join Key Column",
                    source_columns,
                    index=source_columns.index("canonical_item_id") if "canonical_item_id" in source_columns else 0
                )
                source_compare_col = st.selectbox(
                    "Source Text Column to Compare",
                    source_columns,
                    index=0
                )

            with col2:
                target_key_col = st.selectbox(
                    "Target Join Key Column",
                    target_columns,
                    index=target_columns.index("canonical_item_id") if "canonical_item_id" in target_columns else 0
                )
                target_compare_col = st.selectbox(
                    "Target Text Column to Compare",
                    target_columns,
                    index=0
                )

            join_type_label = st.selectbox(
                "Join Type",
                options=[
                    "Keep only IDs found in both files",
                    "Keep all IDs from source file",
                    "Keep all IDs from both files"
                ],
                index=1
            )

            join_type_map = {
                "Keep only IDs found in both files": "inner",
                "Keep all IDs from source file": "left",
                "Keep all IDs from both files": "outer"
            }

            join_type = join_type_map[join_type_label]

        else:
            col1, col2 = st.columns(2)

            with col1:
                source_compare_col = st.selectbox(
                    "Source Column to Compare",
                    source_columns,
                    index=0
                )

            with col2:
                target_compare_col = st.selectbox(
                    "Target Column to Compare",
                    target_columns,
                    index=0
                )

    except Exception as e:
        st.error(f"Error loading workbook metadata: {e}")
else:
    st.info("Upload both files to select sheets and columns.")

run_button = st.button("Run Comparison", type="primary")

if run_button:
    if source_file is None or target_file is None:
        st.error("Please upload both a source file and a target file.")
    else:
        try:
            if comparison_mode.startswith("Key-based"):
                result_df = run_key_based_comparison(
                    comparison_name,
                    source_file,
                    source_sheet,
                    source_key_col,
                    source_compare_col,
                    target_file,
                    target_sheet,
                    target_key_col,
                    target_compare_col,
                    join_type
                )
            else:
                result_df = run_row_by_row_comparison(
                    comparison_name,
                    source_file,
                    source_sheet,
                    source_compare_col,
                    target_file,
                    target_sheet,
                    target_compare_col
                )

            output_file, summary_df, all_results_df = create_output_workbook([result_df])

            st.success("Comparison complete!")

            st.subheader("Summary")
            st.dataframe(summary_df, use_container_width=True)

            st.subheader("Preview Results")
            st.dataframe(all_results_df.head(100), use_container_width=True)

            st.download_button(
                label="Download Results Workbook",
                data=output_file,
                file_name="comparison_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error: {e}")